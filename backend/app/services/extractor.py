"""Document text extraction and token calculation utilities."""

from __future__ import annotations

import csv
import hashlib
import io
import logging
from pathlib import Path
import pypdf
import docx
import openpyxl
import tiktoken

logger = logging.getLogger(__name__)


def compute_sha256(content: bytes) -> str:
    """Compute the SHA-256 hex digest of file content bytes."""
    hasher = hashlib.sha256()
    hasher.update(content)
    return hasher.hexdigest()


def count_tokens(text: str, encoding_name: str = "cl100k_base") -> int:
    """Calculate the number of tokens in a text string using tiktoken."""
    if not text:
        return 0
    try:
        encoding = tiktoken.get_encoding(encoding_name)
        return len(encoding.encode(text))
    except Exception as e:
        logger.warning(f"Failed to encode with tiktoken: {e}. Falling back to rough word estimate.")
        # Fallback heuristic: 1 token ~ 0.75 words (or ~ 4 chars)
        return len(text.split()) + len(text) // 10


def extract_pdf(content: bytes) -> str:
    """Extract plain text from PDF file bytes."""
    text_parts = []
    try:
        pdf_file = io.BytesIO(content)
        reader = pypdf.PdfReader(pdf_file)
        for idx, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                text_parts.append(f"--- Page {idx + 1} ---\n{page_text}")
        return "\n\n".join(text_parts)
    except Exception as e:
        logger.error(f"Error extracting PDF: {e}")
        return f"[PDF Parsing Error: {e}]"


def extract_docx(content: bytes) -> str:
    """Extract plain text from Word document file bytes."""
    try:
        docx_file = io.BytesIO(content)
        doc = docx.Document(docx_file)
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        return "\n".join(text_parts)
    except Exception as e:
        logger.error(f"Error extracting DOCX: {e}")
        return f"[Word Document Parsing Error: {e}]"


def extract_csv(content: bytes) -> str:
    """Extract and format CSV data into a Markdown table."""
    try:
        # Decode as utf-8, fallback to latin-1 if needed
        try:
            csv_text = content.decode("utf-8")
        except UnicodeDecodeError:
            csv_text = content.decode("latin-1")

        csv_reader = csv.reader(io.StringIO(csv_text))
        rows = list(csv_reader)
        if not rows:
            return ""

        output = []
        # Header row
        headers = rows[0]
        output.append("| " + " | ".join(headers) + " |")
        output.append("| " + " | ".join(["---"] * len(headers)) + " |")

        # Data rows
        for row in rows[1:]:
            # Ensure row length matches header length to maintain grid formatting
            padded_row = row + [""] * (len(headers) - len(row))
            # Clean values to avoid issues with pipes in markdown cells
            cleaned_row = [val.replace("|", "\\|").strip() for val in padded_row]
            output.append("| " + " | ".join(cleaned_row) + " |")

        return "\n".join(output)
    except Exception as e:
        logger.error(f"Error extracting CSV: {e}")
        return f"[CSV Parsing Error: {e}]"


def extract_xlsx(content: bytes) -> str:
    """Extract Excel spreadsheet pages and format each sheet as a Markdown table."""
    try:
        xlsx_file = io.BytesIO(content)
        wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
        sheets_markdown = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheets_markdown.append(f"### Sheet: {sheet_name}")
            
            # Read rows
            rows = []
            max_cols = 0
            for r in sheet.iter_rows(values_only=True):
                # Clean none/empty values at the end of rows
                row_vals = [str(val).strip() if val is not None else "" for val in r]
                # Skip entirely empty rows
                if not any(row_vals):
                    continue
                rows.append(row_vals)
                max_cols = max(max_cols, len(row_vals))

            if not rows:
                sheets_markdown.append("*(Empty sheet)*")
                continue

            headers = rows[0]
            # Ensure headers match maximum columns found
            if len(headers) < max_cols:
                headers = headers + [f"Col {i+1}" for i in range(len(headers), max_cols)]

            output = []
            output.append("| " + " | ".join(headers) + " |")
            output.append("| " + " | ".join(["---"] * len(headers)) + " |")

            for row in rows[1:]:
                padded_row = row + [""] * (len(headers) - len(row))
                cleaned_row = [val.replace("|", "\\|") for val in padded_row]
                output.append("| " + " | ".join(cleaned_row) + " |")

            sheets_markdown.append("\n".join(output))

        return "\n\n".join(sheets_markdown)
    except Exception as e:
        logger.error(f"Error extracting XLSX: {e}")
        return f"[Excel Parsing Error: {e}]"


def extract_text(content: bytes) -> str:
    """Extract plain text from general text/markdown file bytes."""
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return content.decode("latin-1")
        except Exception as e:
            logger.error(f"Error decoding text content: {e}")
            return f"[Text Decoding Error: {e}]"


def process_document_content(content: bytes, mime_type: str) -> str:
    """Route document bytes to the correct parser based on MIME type and return plain-text/markdown."""
    mime_type = mime_type.lower()
    if mime_type == "application/pdf":
        return extract_pdf(content)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword"
    ):
        return extract_docx(content)
    elif mime_type == "text/csv":
        return extract_csv(content)
    elif mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel"
    ):
        return extract_xlsx(content)
    elif mime_type.startswith("text/") or mime_type in ("application/json", "application/xml"):
        return extract_text(content)
    else:
        return ""
